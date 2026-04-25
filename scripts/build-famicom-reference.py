from __future__ import annotations

import json
import re
import difflib
from dataclasses import dataclass
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX_PATH = Path.home() / "Downloads" / "Famicom Games checklist.xlsx"
CATALOG_PATH = ROOT / "public" / "catalogs" / "catalog-famicom.json"
OUTPUT_PATH = ROOT / "src" / "famicomReference.ts"


def normalize_search_text(value: str) -> str:
    text = str(value or "").lower()
    text = text.replace("&", " and ")
    text = re.sub(r"\biii\b", " 3 ", text)
    text = re.sub(r"\bii\b", " 2 ", text)
    text = re.sub(r"\biv\b", " 4 ", text)
    text = re.sub(r"\bv\b", " 5 ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def normalize_code(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


@dataclass
class SheetRow:
    title: str
    product_id: str
    release_date: str
    publisher: str


def load_sheet_rows(xlsx_path: Path) -> list[SheetRow]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook["Sheet1"]
    rows: list[SheetRow] = []

    for row_index in range(3, worksheet.max_row + 1):
        title = worksheet.cell(row_index, 1).value
        product_id = worksheet.cell(row_index, 2).value
        release_date = worksheet.cell(row_index, 3).value
        publisher = worksheet.cell(row_index, 4).value

        if not title or not product_id:
            continue

        title_text = str(title).strip()
        product_id_text = str(product_id).strip()

        if title_text in {"A", "Title"} or product_id_text in {"TOP", "ID"}:
            continue

        rows.append(
            SheetRow(
                title=title_text,
                product_id=product_id_text,
                release_date=str(release_date).strip() if release_date else "",
                publisher=str(publisher).strip() if publisher else "",
            )
        )

    return rows


def load_catalog_entries() -> list[dict]:
    return json.loads(CATALOG_PATH.read_text(encoding="utf8"))


def match_rows_to_catalog(rows: list[SheetRow], catalog: list[dict]) -> dict[str, dict]:
    by_normalized_title = {
        normalize_search_text(entry["title"]): entry for entry in catalog if entry.get("id") and entry.get("title")
    }
    catalog_title_tuples = [
        (entry["id"], entry["title"], normalize_search_text(entry["title"])) for entry in catalog if entry.get("id") and entry.get("title")
    ]

    matched: dict[str, dict] = {}

    for row in rows:
        normalized_title = normalize_search_text(row.title)
        entry = by_normalized_title.get(normalized_title)

        if not entry:
            candidates: list[tuple[float, str, str]] = []

            for game_id, title, candidate_normalized in catalog_title_tuples:
                score = difflib.SequenceMatcher(None, normalized_title, candidate_normalized).ratio()
                if score >= 0.84:
                    candidates.append((score, game_id, title))

            candidates.sort(reverse=True)

            if candidates and (len(candidates) == 1 or candidates[0][0] - candidates[1][0] >= 0.08):
                game_id, title = candidates[0][1], candidates[0][2]
                entry = {"id": game_id, "title": title}

        if not entry:
            continue

        game_id = entry["id"]
        alias = row.title if normalize_search_text(row.title) != normalize_search_text(entry["title"]) else ""

        matched[game_id] = {
            "productId": row.product_id,
            "publisher": row.publisher,
            "releaseDate": row.release_date,
            "alias": alias,
        }

    return matched


def render_ts_module(reference_map: dict[str, dict]) -> str:
    lines = [
        "export type FamicomReferenceEntry = {",
        "  productId: string",
        "  publisher?: string",
        "  releaseDate?: string",
        "  alias?: string",
        "}",
        "",
        "export const famicomReferenceByGameId: Record<string, FamicomReferenceEntry> = {",
    ]

    for game_id in sorted(reference_map):
        entry = reference_map[game_id]
        parts = [f"productId: {json.dumps(entry['productId'], ensure_ascii=False)}"]
        if entry.get("publisher"):
            parts.append(f"publisher: {json.dumps(entry['publisher'], ensure_ascii=False)}")
        if entry.get("releaseDate"):
            parts.append(f"releaseDate: {json.dumps(entry['releaseDate'], ensure_ascii=False)}")
        if entry.get("alias"):
            parts.append(f"alias: {json.dumps(entry['alias'], ensure_ascii=False)}")
        lines.append(f"  {json.dumps(game_id)}: {{ {', '.join(parts)} }},")

    lines.append("}")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    rows = load_sheet_rows(DEFAULT_XLSX_PATH)
    catalog = load_catalog_entries()
    matched = match_rows_to_catalog(rows, catalog)
    OUTPUT_PATH.write_text(render_ts_module(matched), encoding="utf8")
    print(f"Wrote {len(matched)} Famicom reference entries to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
